import argparse
import openpyxl
import os
import shutil

# Get path to root of directory containing images
parser = argparse.ArgumentParser(
    prog='move_images',
    description='Moves images from date folders to sample ID folders')
parser.add_argument('-p', '--path', type=str, required=True)
args = parser.parse_args()
root_path = os.path.abspath(args.path)
print('Using root path:', root_path, '\n')
 
# Read the Excel file to get mapping of date folder paths to sample IDs
workbook_name = "DateToSampleID.xlsx"
workbook = openpyxl.load_workbook(os.path.join(root_path, workbook_name))
worksheet = workbook.active

# Keeps track of sources and destinations to alert user if duplication occurs
src_dest_map = {}

# start with row 2 instead of 1 since we don't care about headers, iterate up to last row in sheet
for row in range(2, worksheet.max_row):
  # path where the image is coming from
  src_path = worksheet.cell(row = row, column = 1).value.replace('\\', os.path.sep)
  # Split the path into directory and filename
  directory, filename = os.path.split(src_path)
  # Split the directory into subdirectories
  subdirectories = directory.split(os.path.sep)
  image_date = subdirectories[0]

  # sample ID for the image
  sample_id = worksheet.cell(row = row, column = 2).value

  # Notify user if duplicate is encountered
  src_dest_key = f"{sample_id}_{image_date}"
  if src_dest_key in src_dest_map:
    print(f"Duplicate encountered, data will be overwritten for sample ID: {sample_id} from {image_date} with {filename} (previous value was {src_dest_map[src_dest_key]})")
  else:
    src_dest_map[src_dest_key] = filename

  if os.path.exists(src_path):
    # Use sample ID to determine target dir
    images_dir_name = 'Images_SampleID'
    target_dir = os.path.join(root_path, images_dir_name, sample_id)
    target_path = os.path.join(target_dir, f"{image_date}_{filename}")

    # Make sure the target directory exists before copying
    os.makedirs(target_dir, exist_ok=True)
    shutil.copyfile(os.path.join(root_path, src_path), target_path)
  else:
    print(f"{src_path} does not exist, cannot copy it")